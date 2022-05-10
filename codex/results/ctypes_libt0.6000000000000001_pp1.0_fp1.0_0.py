import ctypes
ctypes.windll.user32.SetProcessDPIAware()

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
import os

def create_excel(data):
    wb = Workbook()
    ws = wb.active

    # 设置字体
    style = Font(name='宋体', size=10)

    # 设置格式
    ws.merge_cells('A1:J1')
    ws.merge_cells('A2:J2')
    ws.merge_cells('A3:J3')
    ws.merge_cells('A4:J4')
    ws.merge_cells('A5:J5')
    ws.merge_cells('A6:J6')
    ws.merge_cells('A7:J7')

    # 写入数据
    ws['A1'] = '欣
