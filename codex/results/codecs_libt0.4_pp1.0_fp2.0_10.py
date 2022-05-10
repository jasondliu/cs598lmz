import codecs
codecs.register(lambda name: codecs.lookup('utf-8') if name == 'cp65001' else None)

import os
import sys
import time
import datetime
import argparse
import csv
import re
import json
import threading
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.cell import get_column_letter

from selenium import webdriver

