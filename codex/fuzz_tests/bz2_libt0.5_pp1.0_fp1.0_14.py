import bz2
bz2.BZ2File('data/sample.csv.bz2')

# gzip
import gzip
gzip.open('data/sample.csv.gz')

# zip
import zipfile
zipfile.ZipFile('data/sample.csv.zip')

# xlsx
import openpyxl
wb = openpyxl.load_workbook('data/sample.xlsx')
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet1')
sheet['A1'].value
sheet['B1'].value
sheet.cell(row=1, column=2).value

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

# pandas
import pandas as pd
xlsx = pd.ExcelFile('data/sample.xlsx')
pd.read_excel(xlsx, 'Sheet1')

frame = pd.read_excel(xlsx, 'Sheet1')
