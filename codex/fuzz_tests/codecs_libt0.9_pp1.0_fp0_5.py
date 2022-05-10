import codecs
codecs.register(STREAM)  # register the StreamReader

from openpyxl import workbook
from openpyxl import load_workbook

def parse_xlsx(file_path=None, sheet_name="Sheet1"):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    rows_max_len = ws.max_row
    cells_max_len = ws.max_column
    res = []
    for row in range(1, rows_max_len+1):
        row_info_dict = {}
        cell_vals = []
        for cell in range(1, cells_max_len+1):
            cell_vals.append(ws.cell(row=row, column=cell).value)
            # print row, cell, ws.cell(row=row, column=cell).value
            row_info_dict['cell_vals'] = cell_vals
        row_info_list = [row_info_dict]
        res += row_info_list
    return res


