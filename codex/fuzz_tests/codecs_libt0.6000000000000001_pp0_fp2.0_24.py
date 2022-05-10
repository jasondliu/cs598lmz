import codecs
codecs.register(lambda name: codecs.lookup('utf-8') if name == 'cp65001' else None)

import os
import sys
import csv
import shutil
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import SQLAlchemyError
from tabulate import tabulate
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.utils import get_column_letter
import logging
import pyodbc
from pyodbc import Error
import re
import xlrd
import xlwt
import xlutils
import win32com.client as win32
import win32com.client
import win32api
import win32gui
import win32con
import win32process
import win32
