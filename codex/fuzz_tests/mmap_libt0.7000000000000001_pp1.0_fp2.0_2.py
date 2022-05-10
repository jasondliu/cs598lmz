import mmap
import csv
import array
import os
from collections import Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Open a file
path = "C:\\Users\\Harshit\\Desktop\\UH\\Machine Learning\\Assignment\\"
dirs = os.listdir(path)
dirs.sort()
dirs.remove('REPORT.xlsx')
dirs.remove('REPORT1.xlsx')
print(dirs)
wb = Workbook()
dest_filename = 'REPORT.xlsx'
ws1 = wb.active
ws1.title = "Sheet1"
ws2 = wb.create_sheet(title="Sheet2")

for folder in dirs:
    print(folder)
    dir = os.listdir(path+folder)
    dir.sort()
    dir.remove('Report.xlsx')
    print(dir)
    cnt = 0
    ws1.cell(row=1, column=1).value = "File Name"
   
